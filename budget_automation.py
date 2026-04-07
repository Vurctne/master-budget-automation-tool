from __future__ import annotations

import csv
import os
import shutil
import tempfile
import time
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator

try:
    import pythoncom
    import win32com.client  # type: ignore
except Exception:
    pythoncom = None
    win32com = None

SUPPORTED_SOURCE_EXTENSIONS = {'.csv', '.xlsx', '.xlsm'}
SUPPORTED_TARGET_EXTENSIONS = {'.xlsx', '.xlsm'}
SUPPORTED_CSV_ENCODINGS = ('utf-8-sig', 'utf-8', 'cp1252')


@dataclass
class ImportSummary:
    output_workbook: str
    report_file: str
    matched_cells: int
    matched_rows: int
    missing_master_codes: List[str]
    missing_master_code_details: List[str]
    missing_source_codes: List[str]
    missing_source_code_details: List[str]
    missing_subprogram_codes: List[str]
    missing_subprogram_details: List[str]
    source_extra_subprogram_codes: List[str]
    source_extra_subprogram_details: List[str]


class BudgetAutomationError(Exception):
    pass


class BudgetAutomator:
    def __init__(self) -> None:
        self.master_sheet_name = 'Master'
        self.compass_sheet_name = 'Compass'
        self.highlight_fill = PatternFill(fill_type='solid', fgColor='FFF4CCCC')
        self.extra_fill = PatternFill(fill_type='solid', fgColor='FFE2F0D9')
        self.protected_start_code = '26201'

    def run(
        self,
        source_file: str,
        template_file: str,
        output_file: str,
        progress_callback: Optional[Callable[[int, str], None]] = None,
    ) -> ImportSummary:
        def progress(percent: int, message: str) -> None:
            if progress_callback is not None:
                progress_callback(percent, message)

        source_path = Path(source_file)
        template_path = Path(template_file)
        output_path = Path(output_file)
        self._validate_paths(source_path, template_path, output_path)
        progress(5, 'Validating files...')

        source_data = self._read_source(source_path)
        progress(15, 'Reading source data...')

        wb = openpyxl.load_workbook(template_path, keep_vba=True, data_only=False)
        try:
            if self.master_sheet_name not in wb.sheetnames:
                raise BudgetAutomationError(f"Sheet '{self.master_sheet_name}' was not found in the workbook.")

            master_ws = wb[self.master_sheet_name]
            protected_start_row = self._find_protected_start_row(master_ws)
            master_map = self._read_master_layout(
                master_ws,
                editable_end_row=self._editable_end_row(master_ws, protected_start_row),
            )
            if not master_map['row_map']:
                raise BudgetAutomationError(
                    'Could not detect the account-code rows on the Master sheet. '
                    'Please check that account codes are in column A.'
                )
            if not master_map['subprogram_map']:
                raise BudgetAutomationError(
                    'Could not detect the sub-program columns on the Master sheet. '
                    'Please check that sub-program codes are in row 4.'
                )
        finally:
            wb.close()

        master_account_codes_for_mismatch = {code for code in master_map['row_codes'] if self._is_mismatch_account_code(code)}
        source_account_codes_for_mismatch = {code for code in source_data['row_codes'] if self._is_mismatch_account_code(code)}
        missing_master_codes = sorted(master_account_codes_for_mismatch - source_account_codes_for_mismatch, key=self._sort_key)
        missing_source_codes = sorted(source_account_codes_for_mismatch - master_account_codes_for_mismatch, key=self._sort_key)
        missing_subprogram_codes = sorted(master_map['subprogram_codes'] - source_data['subprogram_codes'], key=self._sort_key)
        source_extra_subprogram_codes = sorted(source_data['subprogram_codes'] - master_map['subprogram_codes'], key=self._sort_key)

        missing_master_code_details = [
            self._detail_line(code, master_map['row_names'].get(code, '')) for code in missing_master_codes
        ]
        missing_source_code_details = [
            self._detail_line(code, source_data['row_names'].get(code, '')) for code in missing_source_codes
        ]
        missing_subprogram_details = [
            self._detail_line(code, master_map['subprogram_names'].get(code, '')) for code in missing_subprogram_codes
        ]
        source_extra_subprogram_details = [
            self._detail_line(code, source_data['subprogram_names'].get(code, '')) for code in source_extra_subprogram_codes
        ]

        output_path.parent.mkdir(parents=True, exist_ok=True)
        progress(20, 'Checking mismatches and preparing workbook...')

        if self._can_use_excel_native():
            matched_cells, matched_rows = self._run_excel_native(
                template_path=template_path,
                output_path=output_path,
                source_data=source_data,
                missing_master_codes=missing_master_codes,
                missing_source_codes=missing_source_codes,
                missing_subprogram_codes=missing_subprogram_codes,
                source_extra_subprogram_codes=source_extra_subprogram_codes,
                progress_callback=progress,
            )
        else:
            wb = openpyxl.load_workbook(template_path, keep_vba=True, data_only=False)
            try:
                master_ws = wb[self.master_sheet_name]
                compass_ws = wb[self.compass_sheet_name] if self.compass_sheet_name in wb.sheetnames else None
                protected_start_row = self._find_protected_start_row(master_ws)
                progress(35, 'Updating Master sheet...')
                master_map = self._insert_source_only_items_openpyxl(
                    master_ws,
                    source_data,
                    missing_source_codes,
                    source_extra_subprogram_codes,
                    protected_start_row=protected_start_row,
                    progress_callback=progress,
                )
                progress(55, 'Writing imported values...')
                matched_cells, matched_rows = self._populate_master(
                    master_ws,
                    master_map,
                    source_data,
                    progress_callback=progress,
                )
                self._ensure_master_total_formulas_openpyxl(master_ws, master_map)
                if compass_ws is not None:
                    progress(70, 'Refreshing Compass sheet...')
                    self._populate_compass(compass_ws, source_data)
                progress(82, 'Applying mismatch highlighting...')
                self._apply_mismatch_highlights(
                    master_ws,
                    compass_ws,
                    master_map,
                    source_data,
                    missing_master_codes,
                    missing_source_codes,
                    missing_subprogram_codes,
                    source_extra_subprogram_codes,
                )
                progress(92, 'Saving output workbook...')
                wb.save(output_path)
            finally:
                wb.close()

        summary = ImportSummary(
            output_workbook=str(output_path),
            report_file=str(output_path.with_name(output_path.stem + '_import_report.txt')),
            matched_cells=matched_cells,
            matched_rows=matched_rows,
            missing_master_codes=missing_master_codes,
            missing_master_code_details=missing_master_code_details,
            missing_source_codes=missing_source_codes,
            missing_source_code_details=missing_source_code_details,
            missing_subprogram_codes=missing_subprogram_codes,
            missing_subprogram_details=missing_subprogram_details,
            source_extra_subprogram_codes=source_extra_subprogram_codes,
            source_extra_subprogram_details=source_extra_subprogram_details,
        )
        progress(97, 'Writing import report...')
        self._write_report(summary)
        progress(100, 'Completed.')
        return summary

    def _populate_master(
        self,
        ws,
        master_map: Dict[str, Any],
        source_data: Dict[str, Any],
        progress_callback: Optional[Callable[[int, str], None]] = None,
    ) -> tuple[int, int]:
        matched_cells = 0
        matched_rows = 0
        row_items = sorted(master_map['row_map'].items(), key=lambda item: item[1])
        total_rows = max(1, len(row_items))
        for index, (account_code, target_row) in enumerate(row_items, start=1):
            source_row = source_data['rows'].get(account_code)
            if source_row is not None:
                matched_rows += 1
            for subprogram_code, target_col in master_map['subprogram_map'].items():
                cell = ws.cell(target_row, target_col)
                if source_row is None:
                    cell.value = None
                    continue
                source_idx = source_data['subprogram_map'].get(subprogram_code)
                if source_idx is None or source_idx >= len(source_row):
                    cell.value = None
                    continue
                parsed = self._parse_source_number(source_row[source_idx])
                cell.value = parsed
                if parsed is not None:
                    matched_cells += 1
            if progress_callback is not None and (index == total_rows or index % 10 == 0):
                progress = 55 + int(20 * index / total_rows)
                progress_callback(progress, f'Writing imported values... ({index}/{total_rows} rows)')
        return matched_cells, matched_rows

    def _populate_compass(self, ws, source_data: Dict[str, Any]) -> None:
        max_existing_row = ws.max_row
        max_existing_col = ws.max_column
        for row in ws.iter_rows(min_row=1, max_row=max_existing_row, min_col=1, max_col=max_existing_col):
            for cell in row:
                cell.value = None

        ordered_subprogram_codes = [
            code for code in source_data['subprogram_map'].keys() if code not in {'Total', 'EI/SP'}
        ]
        ordered_account_codes = list(source_data['rows'].keys())

        for col_idx, subprogram_code in enumerate(ordered_subprogram_codes, start=4):
            ws.cell(1, col_idx).value = int(subprogram_code) if subprogram_code.isdigit() else subprogram_code
            ws.cell(2, col_idx).value = source_data['subprogram_names'].get(subprogram_code, '')

        ws.cell(2, 2).value = 'EI/SP'
        ws.cell(2, 3).value = 'Total'

        for row_idx, account_code in enumerate(ordered_account_codes, start=3):
            source_row = source_data['rows'][account_code]
            ws.cell(row_idx, 1).value = int(account_code) if account_code.isdigit() else account_code
            ws.cell(row_idx, 2).value = source_data['row_names'].get(account_code, '')
            total_idx = source_data['subprogram_map'].get('Total')
            total_value = self._parse_source_number(source_row[total_idx]) if total_idx is not None and total_idx < len(source_row) else None
            ws.cell(row_idx, 3).value = total_value
            for col_idx, subprogram_code in enumerate(ordered_subprogram_codes, start=4):
                source_idx = source_data['subprogram_map'].get(subprogram_code)
                value = self._parse_source_number(source_row[source_idx]) if source_idx is not None and source_idx < len(source_row) else None
                ws.cell(row_idx, col_idx).value = value

        last_used_row = len(ordered_account_codes) + 2
        last_used_col = len(ordered_subprogram_codes) + 3
        for row in range(1, max_existing_row + 1):
            for col in range(1, max_existing_col + 1):
                if row <= last_used_row and col <= last_used_col:
                    continue
                ws.cell(row, col).value = None

    def _apply_mismatch_highlights(
        self,
        master_ws,
        compass_ws,
        master_map: Dict[str, Any],
        source_data: Dict[str, Any],
        missing_master_codes: List[str],
        missing_source_codes: List[str],
        missing_subprogram_codes: List[str],
        source_extra_subprogram_codes: List[str],
    ) -> None:
        master_last_col = max(master_map['subprogram_map'].values()) if master_map['subprogram_map'] else 3
        master_last_row = max(master_map['row_map'].values()) if master_map['row_map'] else 5
        for code in missing_master_codes:
            row_idx = master_map['row_map'].get(code)
            if row_idx:
                for col in range(1, master_last_col + 1):
                    self._apply_fill(master_ws.cell(row_idx, col))
        for code in missing_subprogram_codes:
            col_idx = master_map['subprogram_map'].get(code)
            if col_idx:
                for row in range(4, master_last_row + 1):
                    self._apply_fill(master_ws.cell(row, col_idx))
        for code in missing_source_codes:
            row_idx = master_map['row_map'].get(code)
            if row_idx:
                for col in range(1, master_last_col + 1):
                    self._apply_extra_fill(master_ws.cell(row_idx, col))
        for code in source_extra_subprogram_codes:
            col_idx = master_map['subprogram_map'].get(code)
            if col_idx:
                for row in range(4, master_last_row + 1):
                    self._apply_extra_fill(master_ws.cell(row, col_idx))

        if compass_ws is None:
            return

        ordered_subprogram_codes = [
            code for code in source_data['subprogram_map'].keys() if code not in {'Total', 'EI/SP'}
        ]
        ordered_account_codes = list(source_data['rows'].keys())
        extra_rows = set(missing_source_codes)
        extra_cols = set(source_extra_subprogram_codes)

        for col_idx, subprogram_code in enumerate(ordered_subprogram_codes, start=4):
            if subprogram_code in extra_cols:
                for row_idx in range(1, len(ordered_account_codes) + 3):
                    self._apply_extra_fill(compass_ws.cell(row_idx, col_idx))

        last_col = len(ordered_subprogram_codes) + 3
        for row_idx, account_code in enumerate(ordered_account_codes, start=3):
            if account_code in extra_rows:
                for col_idx in range(1, last_col + 1):
                    self._apply_extra_fill(compass_ws.cell(row_idx, col_idx))

    def _clear_fill_master(self, ws, master_map: Dict[str, Any]) -> None:
        if not master_map['row_map'] or not master_map['subprogram_map']:
            return
        max_row = ws.max_row
        max_col = ws.max_column
        empty_fill = copy(openpyxl.styles.PatternFill())
        for row in range(4, max_row + 1):
            for col in range(1, max_col + 1):
                ws.cell(row, col).fill = empty_fill

    def _clear_fill_compass(self, ws) -> None:
        empty_fill = copy(openpyxl.styles.PatternFill())
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = empty_fill

    def _can_use_excel_native(self) -> bool:
        return os.name == 'nt' and pythoncom is not None and win32com is not None

    def _run_excel_native(
        self,
        template_path: Path,
        output_path: Path,
        source_data: Dict[str, Any],
        missing_master_codes: List[str],
        missing_source_codes: List[str],
        missing_subprogram_codes: List[str],
        source_extra_subprogram_codes: List[str],
        progress_callback: Optional[Callable[[int, str], None]] = None,
    ) -> tuple[int, int]:
        pythoncom.CoInitialize()
        excel = None
        wb = None
        temp_dir = None
        temp_template_path = None

        def progress(percent: int, message: str) -> None:
            if progress_callback is not None:
                progress_callback(percent, message)

        try:
            if output_path.exists():
                output_path.unlink()

            temp_dir = Path(tempfile.mkdtemp(prefix='budget_automation_'))
            temp_template_path = temp_dir / template_path.name
            shutil.copy2(template_path, temp_template_path)

            excel = win32com.client.DispatchEx('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
            excel.EnableEvents = False
            try:
                excel.AskToUpdateLinks = False
            except Exception:
                pass
            try:
                excel.AutomationSecurity = 3
            except Exception:
                pass

            progress(30, 'Opening workbook in Excel...')
            wb = excel.Workbooks.Open(
                str(temp_template_path.resolve()),
                UpdateLinks=0,
                ReadOnly=False,
                IgnoreReadOnlyRecommended=True,
                AddToMru=False,
                Notify=False,
            )
            master_ws = wb.Worksheets(self.master_sheet_name)
            compass_ws = wb.Worksheets(self.compass_sheet_name) if self.compass_sheet_name in [ws.Name for ws in wb.Worksheets] else None
            protected_start_row = self._find_protected_start_row_excel(master_ws)

            progress(42, 'Updating Master sheet...')
            master_map = self._insert_source_only_items_excel(
                master_ws,
                source_data,
                missing_source_codes,
                source_extra_subprogram_codes,
                protected_start_row=protected_start_row,
                progress_callback=progress,
            )
            progress(58, 'Writing imported values...')
            matched_cells, matched_rows = self._populate_master_excel(
                master_ws,
                master_map,
                source_data,
                progress_callback=progress,
            )
            self._ensure_master_total_formulas_excel(master_ws, master_map)
            if compass_ws is not None:
                progress(72, 'Refreshing Compass sheet...')
                self._populate_compass_excel(compass_ws, source_data)
            progress(84, 'Applying mismatch highlighting...')
            self._apply_mismatch_highlights_excel(
                master_ws,
                compass_ws,
                master_map,
                source_data,
                missing_master_codes,
                missing_source_codes,
                missing_subprogram_codes,
                source_extra_subprogram_codes,
            )

            file_format = 52 if output_path.suffix.lower() == '.xlsm' else 51
            progress(94, 'Saving output workbook...')
            wb.SaveAs(str(output_path.resolve()), FileFormat=file_format, ConflictResolution=2)
            return matched_cells, matched_rows
        except Exception as exc:
            raise BudgetAutomationError(f'Excel save failed: {exc}') from exc
        finally:
            try:
                if wb is not None:
                    wb.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.ScreenUpdating = True
                    excel.EnableEvents = True
                    excel.Quit()
            except Exception:
                pass
            if temp_dir is not None:
                try:
                    shutil.rmtree(temp_dir, ignore_errors=True)
                except Exception:
                    pass
            pythoncom.CoUninitialize()

    def _populate_master_excel(
        self,
        ws,
        master_map: Dict[str, Any],
        source_data: Dict[str, Any],
        progress_callback: Optional[Callable[[int, str], None]] = None,
    ) -> tuple[int, int]:
        sorted_rows = sorted(master_map['row_map'].items(), key=lambda item: item[1])
        sorted_cols = sorted(master_map['subprogram_map'].items(), key=lambda item: item[1])
        if not sorted_rows or not sorted_cols:
            return 0, 0

        matched_cells = 0
        matched_rows = 0
        segments = self._contiguous_row_segments(sorted_rows)
        total_segments = max(1, len(segments))
        for seg_index, segment in enumerate(segments, start=1):
            first_row = segment[0][1]
            last_row = segment[-1][1]
            first_col = sorted_cols[0][1]
            last_col = sorted_cols[-1][1]
            ws.Range(ws.Cells(first_row, first_col), ws.Cells(last_row, last_col)).ClearContents()

            matrix = []
            for account_code, _target_row in segment:
                source_row = source_data['rows'].get(account_code)
                if source_row is not None:
                    matched_rows += 1
                row_values = []
                for subprogram_code, _target_col in sorted_cols:
                    value = ''
                    if source_row is not None:
                        source_idx = source_data['subprogram_map'].get(subprogram_code)
                        if source_idx is not None and source_idx < len(source_row):
                            parsed = self._parse_source_number(source_row[source_idx])
                            value = '' if parsed is None else parsed
                            if parsed is not None:
                                matched_cells += 1
                    row_values.append(value)
                matrix.append(tuple(row_values))

            ws.Range(ws.Cells(first_row, first_col), ws.Cells(last_row, last_col)).Value = tuple(matrix)
            if progress_callback is not None:
                progress = 58 + int(18 * seg_index / total_segments)
                progress_callback(progress, f'Writing imported values... ({seg_index}/{total_segments} sections)')
        return matched_cells, matched_rows

    def _contiguous_row_segments(self, sorted_rows: List[tuple[str, int]]) -> List[List[tuple[str, int]]]:
        if not sorted_rows:
            return []
        segments: List[List[tuple[str, int]]] = [[sorted_rows[0]]]
        for item in sorted_rows[1:]:
            if item[1] == segments[-1][-1][1] + 1:
                segments[-1].append(item)
            else:
                segments.append([item])
        return segments

    def _populate_compass_excel(self, ws, source_data: Dict[str, Any]) -> None:
        max_existing_row = int(ws.UsedRange.Rows.Count)
        max_existing_col = int(ws.UsedRange.Columns.Count)
        ws.Cells.ClearContents()

        ordered_subprogram_codes = [
            code for code in source_data['subprogram_map'].keys() if code not in {'Total', 'EI/SP'}
        ]
        ordered_account_codes = list(source_data['rows'].keys())

        if ordered_subprogram_codes:
            header_row_1 = []
            header_row_2 = []
            for subprogram_code in ordered_subprogram_codes:
                header_row_1.append(int(subprogram_code) if subprogram_code.isdigit() else subprogram_code)
                header_row_2.append(source_data['subprogram_names'].get(subprogram_code, ''))
            ws.Range(ws.Cells(1, 4), ws.Cells(1, 3 + len(header_row_1))).Value = (tuple(header_row_1),)
            ws.Range(ws.Cells(2, 4), ws.Cells(2, 3 + len(header_row_2))).Value = (tuple(header_row_2),)

        ws.Cells(2, 2).Value = 'EI/SP'
        ws.Cells(2, 3).Value = 'Total'

        data_matrix = []
        for account_code in ordered_account_codes:
            source_row = source_data['rows'][account_code]
            total_idx = source_data['subprogram_map'].get('Total')
            total_value = ''
            if total_idx is not None and total_idx < len(source_row):
                parsed_total = self._parse_source_number(source_row[total_idx])
                total_value = '' if parsed_total is None else parsed_total

            row_values = [
                int(account_code) if account_code.isdigit() else account_code,
                source_data['row_names'].get(account_code, ''),
                total_value,
            ]
            for subprogram_code in ordered_subprogram_codes:
                source_idx = source_data['subprogram_map'].get(subprogram_code)
                value = ''
                if source_idx is not None and source_idx < len(source_row):
                    parsed_value = self._parse_source_number(source_row[source_idx])
                    value = '' if parsed_value is None else parsed_value
                row_values.append(value)
            data_matrix.append(tuple(row_values))

        if data_matrix:
            ws.Range(ws.Cells(3, 1), ws.Cells(2 + len(data_matrix), len(data_matrix[0]))).Value = tuple(data_matrix)

        last_used_row = len(ordered_account_codes) + 2
        last_used_col = len(ordered_subprogram_codes) + 3
        if max_existing_row > last_used_row:
            ws.Range(ws.Cells(last_used_row + 1, 1), ws.Cells(max_existing_row, max_existing_col)).ClearContents()
        if max_existing_col > last_used_col and last_used_row >= 1:
            ws.Range(ws.Cells(1, last_used_col + 1), ws.Cells(last_used_row, max_existing_col)).ClearContents()

    def _clear_fill_excel_range(self, ws, row1: int, col1: int, row2: int, col2: int) -> None:
        if ws is None or row2 < row1 or col2 < col1:
            return
        ws.Range(ws.Cells(row1, col1), ws.Cells(row2, col2)).Interior.Pattern = -4142

    def _apply_mismatch_highlights_excel(
        self,
        master_ws,
        compass_ws,
        master_map: Dict[str, Any],
        source_data: Dict[str, Any],
        missing_master_codes: List[str],
        missing_source_codes: List[str],
        missing_subprogram_codes: List[str],
        source_extra_subprogram_codes: List[str],
    ) -> None:
        master_last_col = max(master_map['subprogram_map'].values()) if master_map['subprogram_map'] else 3
        master_last_row = max(master_map['row_map'].values()) if master_map['row_map'] else 5
        for code in missing_master_codes:
            row_idx = master_map['row_map'].get(code)
            if row_idx:
                master_ws.Range(master_ws.Cells(row_idx, 1), master_ws.Cells(row_idx, master_last_col)).Interior.Color = 13421823
        for code in missing_subprogram_codes:
            col_idx = master_map['subprogram_map'].get(code)
            if col_idx:
                master_ws.Range(master_ws.Cells(4, col_idx), master_ws.Cells(master_last_row, col_idx)).Interior.Color = 13421823
        for code in missing_source_codes:
            row_idx = master_map['row_map'].get(code)
            if row_idx:
                master_ws.Range(master_ws.Cells(row_idx, 1), master_ws.Cells(row_idx, master_last_col)).Interior.Color = 14282978
        for code in source_extra_subprogram_codes:
            col_idx = master_map['subprogram_map'].get(code)
            if col_idx:
                master_ws.Range(master_ws.Cells(4, col_idx), master_ws.Cells(master_last_row, col_idx)).Interior.Color = 14282978

        if compass_ws is None:
            return

        ordered_subprogram_codes = [
            code for code in source_data['subprogram_map'].keys() if code not in {'Total', 'EI/SP'}
        ]
        ordered_account_codes = list(source_data['rows'].keys())
        extra_rows = set(missing_source_codes)
        extra_cols = set(source_extra_subprogram_codes)

        for col_idx, subprogram_code in enumerate(ordered_subprogram_codes, start=4):
            if subprogram_code in extra_cols:
                compass_ws.Range(compass_ws.Cells(1, col_idx), compass_ws.Cells(len(ordered_account_codes) + 2, col_idx)).Interior.Color = 14282978

        last_col = len(ordered_subprogram_codes) + 3
        for row_idx, account_code in enumerate(ordered_account_codes, start=3):
            if account_code in extra_rows:
                compass_ws.Range(compass_ws.Cells(row_idx, 1), compass_ws.Cells(row_idx, last_col)).Interior.Color = 14282978

    def _insert_source_only_items_openpyxl(
        self,
        ws,
        source_data: Dict[str, Any],
        missing_source_codes: List[str],
        source_extra_subprogram_codes: List[str],
        protected_start_row: Optional[int] = None,
        progress_callback: Optional[Callable[[int, str], None]] = None,
    ) -> Dict[str, Any]:
        editable_end_row = self._editable_end_row(ws, protected_start_row)
        master_map = self._read_master_layout(ws, editable_end_row=editable_end_row)
        total_steps = max(1, len(source_extra_subprogram_codes) + len(missing_source_codes))
        step_index = 0

        for code in sorted(source_extra_subprogram_codes, key=self._sort_key):
            insert_col = self._find_insert_col(master_map['subprogram_map'], code)
            source_col = insert_col - 1 if insert_col > 4 else insert_col + 1
            self._insert_partial_column_openpyxl(ws, insert_col, 1, editable_end_row)
            self._copy_column_format_openpyxl(ws, source_col, insert_col)
            self._copy_column_formulas_openpyxl(ws, source_col, insert_col, 1, editable_end_row)
            ws.cell(4, insert_col).value = int(code) if code.isdigit() else code
            ws.cell(5, insert_col).value = source_data['subprogram_names'].get(code, '')
            master_map = self._read_master_layout(
                ws,
                editable_end_row=self._editable_end_row(ws, protected_start_row),
            )
            step_index += 1
            if progress_callback is not None:
                progress = 35 + int(18 * step_index / total_steps)
                progress_callback(progress, f'Updating Master sheet... ({step_index}/{total_steps})')

        for code in sorted(missing_source_codes, key=self._sort_key):
            insert_row = self._find_insert_row(master_map['row_map'], code)
            if protected_start_row is not None and insert_row >= protected_start_row:
                continue
            source_row = insert_row - 1 if insert_row > 6 else insert_row + 1
            ws.insert_rows(insert_row, 1)
            self._copy_row_format_openpyxl(ws, source_row, insert_row)
            self._copy_row_formulas_openpyxl(ws, source_row, insert_row)
            ws.cell(insert_row, 1).value = int(code) if code.isdigit() else code
            ws.cell(insert_row, 2).value = source_data['row_names'].get(code, '')
            master_map = self._read_master_layout(
                ws,
                editable_end_row=self._editable_end_row(ws, protected_start_row),
            )
            self._set_row_total_formula_openpyxl(ws, insert_row, max(master_map['subprogram_map'].values()))
            step_index += 1
            if progress_callback is not None:
                progress = 35 + int(18 * step_index / total_steps)
                progress_callback(progress, f'Updating Master sheet... ({step_index}/{total_steps})')

        return self._read_master_layout(ws, editable_end_row=self._editable_end_row(ws, protected_start_row))

    def _insert_source_only_items_excel(
        self,
        ws,
        source_data: Dict[str, Any],
        missing_source_codes: List[str],
        source_extra_subprogram_codes: List[str],
        protected_start_row: Optional[int] = None,
        progress_callback: Optional[Callable[[int, str], None]] = None,
    ) -> Dict[str, Any]:
        editable_end_row = self._editable_end_row_excel(ws, protected_start_row)
        master_map = self._read_master_layout_excel(ws, editable_end_row=editable_end_row)
        total_steps = max(1, len(source_extra_subprogram_codes) + len(missing_source_codes))
        step_index = 0

        for code in sorted(source_extra_subprogram_codes, key=self._sort_key):
            insert_col = self._find_insert_col(master_map['subprogram_map'], code)
            copy_from_col = insert_col - 1 if insert_col > 4 else insert_col + 1
            self._insert_partial_column_excel(ws, insert_col, 1, editable_end_row)
            try:
                ws.Columns(insert_col).ColumnWidth = ws.Columns(copy_from_col).ColumnWidth
            except Exception:
                pass
            self._copy_column_formulas_excel(ws, copy_from_col, insert_col, 1, editable_end_row)
            ws.Cells(4, insert_col).Value = int(code) if code.isdigit() else code
            ws.Cells(5, insert_col).Value = source_data['subprogram_names'].get(code, '')
            master_map = self._read_master_layout_excel(
                ws,
                editable_end_row=self._editable_end_row_excel(ws, protected_start_row),
            )
            step_index += 1
            if progress_callback is not None:
                progress = 42 + int(14 * step_index / total_steps)
                progress_callback(progress, f'Updating Master sheet... ({step_index}/{total_steps})')

        for code in sorted(missing_source_codes, key=self._sort_key):
            insert_row = self._find_insert_row(master_map['row_map'], code)
            if protected_start_row is not None and insert_row >= protected_start_row:
                continue
            copy_from_row = insert_row - 1 if insert_row > 6 else insert_row + 1
            ws.Rows(insert_row).Insert()
            try:
                ws.Rows(insert_row).RowHeight = ws.Rows(copy_from_row).RowHeight
            except Exception:
                pass
            self._copy_row_formulas_excel(ws, copy_from_row, insert_row, int(ws.UsedRange.Columns.Count))
            ws.Cells(insert_row, 1).Value = int(code) if code.isdigit() else code
            ws.Cells(insert_row, 2).Value = source_data['row_names'].get(code, '')
            master_map = self._read_master_layout_excel(
                ws,
                editable_end_row=self._editable_end_row_excel(ws, protected_start_row),
            )
            self._set_row_total_formula_excel(ws, insert_row, max(master_map['subprogram_map'].values()))
            step_index += 1
            if progress_callback is not None:
                progress = 42 + int(14 * step_index / total_steps)
                progress_callback(progress, f'Updating Master sheet... ({step_index}/{total_steps})')

        return self._read_master_layout_excel(ws, editable_end_row=self._editable_end_row_excel(ws, protected_start_row))

    def _insert_partial_column_openpyxl(self, ws, insert_col: int, start_row: int, end_row: int) -> None:
        if end_row < start_row:
            return
        max_col_before = ws.max_column
        ws.move_range(
            f'{get_column_letter(insert_col)}{start_row}:{get_column_letter(max_col_before)}{end_row}',
            rows=0,
            cols=1,
            translate=True,
        )
        for row in range(start_row, end_row + 1):
            ws.cell(row, insert_col).value = None

    def _copy_column_formulas_openpyxl(self, ws, source_col: int, target_col: int, start_row: int, end_row: int) -> None:
        if source_col < 1 or end_row < start_row:
            return
        for row in range(start_row, end_row + 1):
            source_value = ws.cell(row, source_col).value
            if isinstance(source_value, str) and source_value.startswith('='):
                try:
                    translated = Translator(
                        source_value,
                        origin=f'{get_column_letter(source_col)}{row}',
                    ).translate_formula(f'{get_column_letter(target_col)}{row}')
                except Exception:
                    translated = source_value
                ws.cell(row, target_col).value = translated

    def _copy_row_formulas_openpyxl(self, ws, source_row: int, target_row: int) -> None:
        if source_row < 1 or source_row > ws.max_row:
            return
        for col in range(1, ws.max_column + 1):
            source_value = ws.cell(source_row, col).value
            if isinstance(source_value, str) and source_value.startswith('='):
                try:
                    translated = Translator(
                        source_value,
                        origin=f'{get_column_letter(col)}{source_row}',
                    ).translate_formula(f'{get_column_letter(col)}{target_row}')
                except Exception:
                    translated = source_value
                ws.cell(target_row, col).value = translated

    def _insert_partial_column_excel(self, ws, insert_col: int, start_row: int, end_row: int) -> None:
        if end_row < start_row:
            return
        ws.Range(ws.Cells(start_row, insert_col), ws.Cells(end_row, insert_col)).Insert(Shift=-4161)

    def _copy_column_formulas_excel(self, ws, source_col: int, target_col: int, start_row: int, end_row: int) -> None:
        if source_col < 1 or end_row < start_row:
            return
        for row in range(start_row, end_row + 1):
            try:
                if bool(ws.Cells(row, source_col).HasFormula):
                    ws.Cells(row, target_col).FormulaR1C1 = ws.Cells(row, source_col).FormulaR1C1
            except Exception:
                pass

    def _copy_row_formulas_excel(self, ws, source_row: int, target_row: int, end_col: int) -> None:
        if source_row < 1:
            return
        for col in range(1, int(end_col) + 1):
            try:
                if bool(ws.Cells(source_row, col).HasFormula):
                    ws.Cells(target_row, col).FormulaR1C1 = ws.Cells(source_row, col).FormulaR1C1
            except Exception:
                pass

    def _copy_column_format_openpyxl(self, ws, source_col: int, target_col: int) -> None:
        if source_col < 1 or source_col > ws.max_column:
            return
        for row in range(1, ws.max_row + 1):
            ws.cell(row, target_col)._style = copy(ws.cell(row, source_col)._style)
            if ws.row_dimensions[row].height is not None:
                ws.row_dimensions[row].height = ws.row_dimensions[row].height
        source_letter = get_column_letter(source_col)
        target_letter = get_column_letter(target_col)
        ws.column_dimensions[target_letter].width = ws.column_dimensions[source_letter].width
        ws.column_dimensions[target_letter].hidden = ws.column_dimensions[source_letter].hidden

    def _copy_row_format_openpyxl(self, ws, source_row: int, target_row: int) -> None:
        if source_row < 1 or source_row > ws.max_row:
            return
        for col in range(1, ws.max_column + 1):
            ws.cell(target_row, col)._style = copy(ws.cell(source_row, col)._style)
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
        ws.row_dimensions[target_row].hidden = ws.row_dimensions[source_row].hidden

    def _copy_excel_column_format(self, ws, source_col: int, target_col: int, excel) -> None:
        if source_col < 1:
            return
        ws.Columns(source_col).Copy()
        ws.Columns(target_col).PasteSpecial(Paste=-4122)
        try:
            ws.Columns(target_col).ColumnWidth = ws.Columns(source_col).ColumnWidth
        except Exception:
            pass
        excel.CutCopyMode = False

    def _copy_excel_row_format(self, ws, source_row: int, target_row: int, excel) -> None:
        if source_row < 1:
            return
        ws.Rows(source_row).Copy()
        ws.Rows(target_row).PasteSpecial(Paste=-4122)
        try:
            ws.Rows(target_row).RowHeight = ws.Rows(source_row).RowHeight
        except Exception:
            pass
        excel.CutCopyMode = False

    def _ensure_master_total_formulas_openpyxl(self, ws, master_map: Dict[str, Any]) -> None:
        if not master_map['row_map'] or not master_map['subprogram_map']:
            return
        last_col = max(master_map['subprogram_map'].values())
        for row_idx in master_map['row_map'].values():
            self._set_row_total_formula_openpyxl(ws, row_idx, last_col)

    def _ensure_master_total_formulas_excel(self, ws, master_map: Dict[str, Any]) -> None:
        if not master_map['row_map'] or not master_map['subprogram_map']:
            return
        last_col = max(master_map['subprogram_map'].values())
        for row_idx in master_map['row_map'].values():
            self._set_row_total_formula_excel(ws, row_idx, last_col)

    def _set_row_total_formula_openpyxl(self, ws, row_idx: int, last_col: int) -> None:
        last_col_letter = get_column_letter(last_col)
        ws.cell(row_idx, 3).value = f'=SUM(D{row_idx}:{last_col_letter}{row_idx})'

    def _set_row_total_formula_excel(self, ws, row_idx: int, last_col: int) -> None:
        last_col_letter = get_column_letter(last_col)
        ws.Cells(row_idx, 3).Formula = f'=SUM(D{row_idx}:{last_col_letter}{row_idx})'

    def _find_insert_col(self, subprogram_map: Dict[str, int], new_code: str) -> int:
        ordered = sorted(subprogram_map.items(), key=lambda item: item[1])
        for code, col in ordered:
            if self._sort_key(new_code) < self._sort_key(code):
                return col
        return (ordered[-1][1] + 1) if ordered else 4

    def _find_insert_row(self, row_map: Dict[str, int], new_code: str) -> int:
        ordered = sorted(row_map.items(), key=lambda item: item[1])
        for code, row in ordered:
            if self._sort_key(new_code) < self._sort_key(code):
                return row
        return (ordered[-1][1] + 1) if ordered else 6

    def _validate_paths(self, source_path: Path, template_path: Path, output_path: Path) -> None:
        source_resolved = source_path.resolve()
        template_resolved = template_path.resolve()
        output_resolved = output_path.resolve()
        if not source_path.exists():
            raise BudgetAutomationError(f'Source file not found: {source_path}')
        if not template_path.exists():
            raise BudgetAutomationError(f'Target workbook not found: {template_path}')
        if source_path.suffix.lower() not in SUPPORTED_SOURCE_EXTENSIONS:
            raise BudgetAutomationError('Source file must be one of: .csv, .xlsx, .xlsm')
        if template_path.suffix.lower() not in SUPPORTED_TARGET_EXTENSIONS:
            raise BudgetAutomationError('Target workbook must be one of: .xlsx, .xlsm')
        if output_path.suffix.lower() not in SUPPORTED_TARGET_EXTENSIONS:
            raise BudgetAutomationError('Output workbook must be one of: .xlsx, .xlsm')
        if source_resolved == template_resolved:
            raise BudgetAutomationError('Source file and target workbook must be different files.')
        if source_resolved == output_resolved:
            raise BudgetAutomationError('Output workbook must be different from the source file.')
        if template_resolved == output_resolved:
            raise BudgetAutomationError('Output workbook must be a new file. Please do not save over the original template.')
        if template_path.suffix.lower() == '.xlsm' and output_path.suffix.lower() != '.xlsm':
            raise BudgetAutomationError(
                'Output workbook must use the .xlsm extension when the template workbook is .xlsm, '
                'otherwise macros and button bindings cannot be preserved.'
            )

    def _read_source(self, source_path: Path) -> Dict[str, Any]:
        rows = self._read_csv_rows(source_path) if source_path.suffix.lower() == '.csv' else self._read_excel_rows(source_path)
        if len(rows) < 3:
            raise BudgetAutomationError('Source file does not contain enough rows.')

        normalized = self._normalize_rows(rows)
        subprogram_row = normalized[0]
        name_row = normalized[1]

        subprogram_map: Dict[str, int] = {}
        subprogram_names: Dict[str, str] = {}
        duplicate_subprogram_codes: List[str] = []
        for idx, code in enumerate(subprogram_row):
            code_clean = self._clean_string(code)
            if code_clean:
                if code_clean in subprogram_map:
                    duplicate_subprogram_codes.append(code_clean)
                    continue
                subprogram_map[code_clean] = idx
                subprogram_names[code_clean] = self._clean_string(name_row[idx]) if idx < len(name_row) else ''

        data_rows: Dict[str, List[str]] = {}
        row_names: Dict[str, str] = {}
        duplicate_account_codes: List[str] = []
        for row in normalized[2:]:
            account_code = self._clean_string(row[0])
            if not account_code:
                continue
            if account_code in data_rows:
                duplicate_account_codes.append(account_code)
                continue
            data_rows[account_code] = row
            row_names[account_code] = self._clean_string(row[1]) if len(row) > 1 else ''

        if duplicate_subprogram_codes:
            duplicates = ', '.join(sorted(set(duplicate_subprogram_codes), key=self._sort_key))
            raise BudgetAutomationError(f'Source file contains duplicate sub-program codes: {duplicates}')
        if duplicate_account_codes:
            duplicates = ', '.join(sorted(set(duplicate_account_codes), key=self._sort_key))
            raise BudgetAutomationError(f'Source file contains duplicate account codes: {duplicates}')

        return {
            'rows': data_rows,
            'row_names': row_names,
            'row_codes': set(data_rows.keys()),
            'subprogram_map': subprogram_map,
            'subprogram_names': subprogram_names,
            'subprogram_codes': set(subprogram_map.keys()) - {'Total', 'EI/SP'},
        }

    def _read_csv_rows(self, source_path: Path) -> List[List[str]]:
        last_error: UnicodeDecodeError | None = None
        for encoding in SUPPORTED_CSV_ENCODINGS:
            try:
                with source_path.open('r', encoding=encoding, newline='') as handle:
                    return list(csv.reader(handle))
            except UnicodeDecodeError as exc:
                last_error = exc
        raise BudgetAutomationError(
            'Could not read the CSV source file using UTF-8 or Windows encodings. '
            'Please re-save the file as UTF-8 CSV and try again.'
        ) from last_error

    def _read_excel_rows(self, source_path: Path) -> List[List[str]]:
        wb = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(['' if value is None else str(value) for value in row])
            return rows
        finally:
            wb.close()

    def _normalize_rows(self, rows: List[List[str]]) -> List[List[str]]:
        non_empty = [row for row in rows if any(str(cell).strip() for cell in row)]
        if not non_empty:
            raise BudgetAutomationError('Source file is empty.')
        max_len = max(len(row) for row in non_empty)
        return [row + [''] * (max_len - len(row)) for row in non_empty]

    def _find_protected_start_row(self, ws) -> Optional[int]:
        for row in range(1, ws.max_row + 1):
            if self._clean_string(ws.cell(row, 1).value) == self.protected_start_code:
                return row
        return None

    def _find_protected_start_row_excel(self, ws) -> Optional[int]:
        max_row = int(ws.UsedRange.Rows.Count)
        for row in range(1, max_row + 1):
            if self._clean_string(ws.Cells(row, 1).Value) == self.protected_start_code:
                return row
        return None

    @staticmethod
    def _editable_end_row(ws, protected_start_row: Optional[int]) -> int:
        return (protected_start_row - 1) if protected_start_row else ws.max_row

    @staticmethod
    def _editable_end_row_excel(ws, protected_start_row: Optional[int]) -> int:
        max_row = int(ws.UsedRange.Rows.Count)
        return (protected_start_row - 1) if protected_start_row else max_row

    def _read_master_layout(self, ws, editable_end_row: Optional[int] = None) -> Dict[str, Any]:
        row_map: Dict[str, int] = {}
        row_names: Dict[str, str] = {}
        subprogram_map: Dict[str, int] = {}
        subprogram_names: Dict[str, str] = {}
        for col in range(4, ws.max_column + 1):
            code = self._clean_string(ws.cell(4, col).value)
            if code:
                subprogram_map[code] = col
                subprogram_names[code] = self._clean_string(ws.cell(5, col).value)
        max_row = editable_end_row if editable_end_row is not None else ws.max_row
        for row in range(1, max_row + 1):
            account_code = self._clean_string(ws.cell(row, 1).value)
            if account_code.isdigit() and len(account_code) >= 5:
                row_map[account_code] = row
                row_names[account_code] = self._clean_string(ws.cell(row, 2).value)
        return {
            'row_map': row_map,
            'row_names': row_names,
            'subprogram_map': subprogram_map,
            'subprogram_names': subprogram_names,
            'row_codes': set(row_map.keys()),
            'subprogram_codes': set(subprogram_map.keys()),
        }

    def _read_master_layout_excel(self, ws, editable_end_row: Optional[int] = None) -> Dict[str, Any]:
        row_map: Dict[str, int] = {}
        row_names: Dict[str, str] = {}
        subprogram_map: Dict[str, int] = {}
        subprogram_names: Dict[str, str] = {}
        max_col = int(ws.UsedRange.Columns.Count)
        max_row = editable_end_row if editable_end_row is not None else int(ws.UsedRange.Rows.Count)
        for col in range(4, max_col + 1):
            code = self._clean_string(ws.Cells(4, col).Value)
            if code:
                subprogram_map[code] = col
                subprogram_names[code] = self._clean_string(ws.Cells(5, col).Value)
        for row in range(1, max_row + 1):
            account_code = self._clean_string(ws.Cells(row, 1).Value)
            if account_code.isdigit() and len(account_code) >= 5:
                row_map[account_code] = row
                row_names[account_code] = self._clean_string(ws.Cells(row, 2).Value)
        return {
            'row_map': row_map,
            'row_names': row_names,
            'subprogram_map': subprogram_map,
            'subprogram_names': subprogram_names,
            'row_codes': set(row_map.keys()),
            'subprogram_codes': set(subprogram_map.keys()),
        }

    def _write_report(self, summary: ImportSummary) -> None:
        mismatch_count = (
            len(summary.missing_master_codes)
            + len(summary.missing_source_codes)
            + len(summary.missing_subprogram_codes)
            + len(summary.source_extra_subprogram_codes)
        )
        overall_status = 'No mismatch or concern found.' if mismatch_count == 0 else f'Mismatch detected: {mismatch_count} item(s) need review.'
        lines = [
            'Budget import completed successfully.',
            overall_status,
            '',
            f'Output workbook: {summary.output_workbook}',
            f'Report file: {summary.report_file}',
            f'Matched rows: {summary.matched_rows}',
            f'Matched cells: {summary.matched_cells}',
            '',
            'Master account codes missing from source:',
            'Description: Master contains a tracked 5-digit account code (starting with 7 or 8) that was not found in the source file.',
            *self._report_lines(summary.missing_master_code_details),
            '',
            'Source account codes not used by Master:',
            'Description: Source contains a tracked 5-digit account code (starting with 7 or 8) that does not exist on the original Master layout. It is added into the Master sheet in numeric position, with description and light green highlighting.',
            *self._report_lines(summary.missing_source_code_details),
            '',
            'Master sub-program codes missing from source:',
            'Description: Master contains a sub-program code that was not found in the source file.',
            *self._report_lines(summary.missing_subprogram_details),
            '',
            'Source sub-program codes not used by Master:',
            'Description: Source contains a sub-program code that does not exist on the original Master layout. It is added into the Master sheet in numeric position, with description and light green highlighting.',
            *self._report_lines(summary.source_extra_subprogram_details),
        ]
        Path(summary.report_file).write_text('\n'.join(lines), encoding='utf-8')

    @staticmethod
    def _report_lines(items: List[str]) -> List[str]:
        return items if items else ['None']

    @staticmethod
    def _detail_line(code: str, description: str) -> str:
        return f'{code} - {description}'.rstrip(' -')

    @staticmethod
    def _is_mismatch_account_code(code: str) -> bool:
        return code.isdigit() and len(code) == 5 and code[0] in {'7', '8'}

    @staticmethod
    def _clean_string(value: Any) -> str:
        if value is None:
            return ''
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def _sort_key(code: str) -> tuple[int, Any]:
        text = str(code).strip()
        if text.isdigit():
            return (0, int(text))
        return (1, text)

    @staticmethod
    def _parse_source_number(value: Any) -> Any:
        if value is None:
            return None
        text = str(value).strip()
        if text == '':
            return None
        if text in {'#N/A', '#VALUE!', '#REF!', '#DIV/0!', '#NAME?', '#NUM!', '#NULL!'}:
            return None
        text = text.replace(',', '')
        try:
            number = float(text)
            return int(number) if number.is_integer() else number
        except ValueError:
            return text

    def _apply_fill(self, cell) -> None:
        cell.fill = copy(self.highlight_fill)

    def _apply_extra_fill(self, cell) -> None:
        cell.fill = copy(self.extra_fill)
